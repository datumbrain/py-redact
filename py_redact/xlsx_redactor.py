import re
import logging
import openpyxl
from openpyxl.utils.cell import get_column_letter

logger = logging.getLogger("docx-redact")


class XlsxRedactor:

    def __init__(self, xlsx_obj_path, regexes, replace_char):
        self.xlsx_obj_path = xlsx_obj_path
        self.regexes = regexes
        self.replace_char = replace_char

    def redact(self, output_file_path):
        """
        Redacts the given .docx file and writes result to output file
        :param output_file_path (string): path of the file to write the result to
        :return:
        """
        wb = openpyxl.load_workbook(self.xlsx_obj_path)
        sheets = wb.sheetnames
        for reg in self.regexes:
            regex = re.compile(reg)
            for sheet_name in sheets:
                sheet = wb[sheet_name]
                number_rows = sheet.max_row
                number_columns = sheet.max_column
                for i in range(number_columns):
                    for k in range(number_rows):
                        cell = sheet[get_column_letter(i + 1) + str(k + 1)]
                        old_val = str(cell.value)
                        match_obj = regex.search(old_val)
                        if match_obj:
                            new_val = regex.sub(self.replace_char * len(match_obj.group(0)), old_val)
                            cell.value = new_val
        if self.xlsx_obj_path == output_file_path:
            logger.warning("Input and Output files are same!")
        wb.save(output_file_path)
        logging.info("Updated file saved as: " + output_file_path)
